import cv2
import os
import time
import numpy as np
from pyzbar.pyzbar import decode
from openpyxl import Workbook
import openpyxl
import subprocess

# Get the directory where the script is located
script_directory = os.path.dirname(os.path.abspath(__file__))
excel_file_path = os.path.join(script_directory, "qr_codes.xlsx")

# Initialize the webcam
cap = cv2.VideoCapture(0)

# Cooldown time (in seconds)
cooldown_time = 1
last_detection_time = time.time()

# Function to close Excel (if it's running)
def close_excel():
    print("Closing Excel if open...")
    try:
        os.system("taskkill /f /im excel.exe")  # For Windows, forcefully close Excel
        time.sleep(2)  # Wait for Excel to fully close
    except Exception as e:
        print(f"Error closing Excel: {e}")

# Function to reopen Excel
def reopen_excel(filename):
    print(f"Reopening Excel with {filename}...")
    try:
        subprocess.Popen(["start", "excel", filename], shell=True)
    except Exception as e:
        print(f"Error reopening Excel: {e}")

# Function to save QR code data to Excel
def append_qr_code_to_excel(row, filename=excel_file_path):
    def save():
        while True:
            try:
                # Check if the file exists
                if not os.path.exists(filename):
                    # Create a new workbook if it doesn't exist
                    workbook = Workbook()
                    sheet = workbook.active
                    sheet.title = "qr_codes"  # Set the sheet name to "qr_codes"
                    sheet.append(["Timestamp", "QR Data"])  # Add headers
                else:
                    # Load the existing workbook
                    workbook = openpyxl.load_workbook(filename)
                    # Check if the 'qr_codes' sheet exists, if not create it
                    if "qr_codes" in workbook.sheetnames:
                        sheet = workbook["qr_codes"]
                    else:
                        sheet = workbook.create_sheet("qr_codes")
                        sheet.append(["Timestamp", "QR Data"])  # Add headers for the new sheet

                # Append the new row to the sheet
                sheet.append(row)
                
                # Save the workbook
                workbook.save(filename)
                print(f"Saved to Excel: {row}")
                break  # Exit the loop once saved successfully
            except PermissionError:
                print(f"Permission denied: {filename}. Retrying in 1 second...")
                time.sleep(1)  # Wait 1 second before retrying
            except Exception as e:
                print(f"Unexpected error while saving to Excel: {e}")
                break

    # Start the save operation
    save()

print("Starting QR code scanner... Press 'q' to quit.")

# Start scanning
while True:
    # Capture frame-by-frame
    ret, frame = cap.read()
    if not ret:
        break
    
    # Decode QR codes in the current frame
    qr_codes = decode(frame)
    
    for qr in qr_codes:
        qr_data = qr.data.decode('utf-8')
        timestamp = time.strftime('%Y-%m-%d %H:%M:%S')  # Get current timestamp
        
        # Check cooldown to avoid multiple detections in quick succession
        current_time = time.time()
        if current_time - last_detection_time > cooldown_time:
            print(f"QR Code detected: {qr_data} at {timestamp}")
            
            # Check if the QR data contains "No Card"
            if "No Card" in qr_data:
                qr_data_parts = qr_data.replace("No Card", "No_Card").split()  # Temporarily replace "No Card" with "No_Card"
                qr_data_parts = [part.replace("No_Card", "No Card") for part in qr_data_parts]  # Restore "No Card"
            else:
                # Split the QR data into parts (assuming tab or space delimiter)
                qr_data_parts = qr_data.split()  # You can use .split('\t') if tabs are used
            
            # Add timestamp as the first column
            qr_data_parts.insert(0, timestamp)
            
            # Close Excel before saving data
            close_excel()

            # Save the processed QR code data immediately
            append_qr_code_to_excel(qr_data_parts)
            
            # Reopen the Excel file after saving
            reopen_excel(excel_file_path)

            last_detection_time = current_time  # Update last detection time

        # Draw the bounding box around the QR code
        rect_points = qr.polygon
        if len(rect_points) == 4:
            pts = [tuple(pt) for pt in rect_points]
            cv2.polylines(frame, [np.array(pts, dtype=np.int32)], isClosed=True, color=(0, 255, 0), thickness=3)
    
    # Show the current frame with detected QR codes
    cv2.imshow("QR Code Scanner", frame)

    # Exit condition
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

# Release the webcam and close the window
cap.release()
cv2.destroyAllWindows()
